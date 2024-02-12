from tkinter import *
from datetime import date
from tkinter import filedialog
from tkinter import messagebox
from PIL import Image, ImageTk
import os
from tkinter.ttk import Combobox
import openpyxl , xlrd
from openpyxl import Workbook
import pathlib
from datetime import datetime
import re
from customtkinter import *
set_default_color_theme("blue")
set_appearance_mode("light")
background = "#06283D"
framebg = "#EDEDED"
framefg = "#06283D"
undo_history = []
redo_history = []
root = CTk()
root.title("Student Registration")
root.geometry("1250x700+210+100")
root.configure(bg=background)

file=pathlib.Path('Students_data.xlsx')
if file.exists():
    pass
else:
    file=Workbook()
    sheet=file.active
    sheet['A1'] = "Registration No."
    sheet['B1'] = "Name"
    sheet['C1'] = "Course"
    sheet['D1'] = "Gender"
    sheet['E1'] = "DOB"
    sheet['F1'] = "Date of Registration"
    sheet['G1'] = "Religion"
    sheet['H1'] = "CNIC"
    sheet['I1'] = "Father's Name"
    sheet['J1'] = "Mother's Name"
    sheet['K1'] = "Father's Occupation"
    sheet['L1'] = "Mother's Occupation"

    file.save('Students_data.xlsx')

######################################----Exit-----Window-----######################################################
def Exit(*args):
    root.destroy()

#############################Upload show image #################################################################
def showimage(*args):
    global filename
    global img
    filename=filedialog.askopenfilename(initialdir=os.getcwd(),
                                        title="Select image file", filetype=(("JPG File", "*.jpg"),
                                                                             ("PNG File", "*.png"),
                                                                             ("All Files", "*.txt")))
    img = (Image.open(filename))
    resized_image=img.resize((190,190))
    photo2 = ImageTk.PhotoImage(resized_image)
    lbl.configure(image=photo2)
    lbl.image=photo2


def Clear(*args):
    global undo_history
    global img
    # Save the current state to undo history
    # undo_history.append(get_all_entry_values())
    
    # Clear all entry fields
    #Registration.set("")
    Name.set("")
    DOB.set("")
    Religion.set("")
    CNIC.set("")
    Date.set(d1)
    Course.set("Select Course")
    F_Name.set("")
    Father_Occupation.set("")
    M_Name.set("")
    Mother_Occupation.set("")
   
    registration_no()
    saveButton.configure(state="normal")


    img1=PhotoImage(file='Images/upload photo.png')
    lbl.configure(image=img1)
    lbl.image=img1
   
    img=""
    # Clear the image label
    img = Image.open("examples/Images/upload photo.png")
    resized_image = img.resize((190, 190))
    photo2 = ImageTk.PhotoImage(resized_image)
    lbl.config(image=photo2)
    lbl.image = photo2

    # Show success message
    messagebox.showinfo("Success", "Data cleared successfully.")
#############################################  REGISTRATION NO(Automation)#########################################################333
def registration_no():
    file=openpyxl.load_workbook('student_data.xlsx')
    sheet=file.active
    row=sheet.row_max
    
    max_row_value=sheet.cell(row=row,column=1).value

    try:
        Registration.set(max_row_value+1)
    except:
        Registration.set("1")    
######################----gender###############################################33
def selection(*args):
    global gender
    value=radio.get()
    if value==1:
        gender="Male"

    else:
        gender="Female"

# def reset_form(*args):
#     Registration.set("")
#     Date.set(d1)
#     Name.set("")
#     DOB.set("")
#     radio.set(0)
#     Religion.set("")
#     Skill.set("")
#     Class.set("Select Class")
#     F_Name.set("")
#     Father_Occupation.set("")
#     M_Name.set("")
#     Mother_Occupation.set("")

def save_data(*args):
    R1=Registration.get()
    N1=Name.get()
    C1=Course.get()
    try:
        G1=gender
    except:
        messagebox.showerror("error","Select Gender")
    D2=DOB.get()
    D1=Date.get()
    Re1=Religion.get()
    S1=CNIC.get()
    fathername=F_Name.get()
    mothername=M_Name.get()
    F1=Father_Occupation.get()
    M1=Mother_Occupation.get()

    if N1=="" or C1=="Select Course" or D2==""  or Re1=="" or S1="" or fathername=="" or mothername=="" or F1=="":
        messaebox.showerror("error","Some Data is missing")
    else:
          file=openpyxl.load_workbook('Student_data.xlsx')
          sheet=file.active
          sheet.cell(row=sheet.max_row+1, column=1, value=R1)
          sheet.cell(row=sheet.max_row, column=2, value=N1)
          sheet.cell(row=sheet.max_row, column=3, value=C1)
          sheet.cell(row=sheet.max_row, column=4, value=G1)
          sheet.cell(row=sheet.max_row,column=5, value=D2)
          sheet.cell(row=sheet.max_row,, column=6, value=D1)
          sheet.cell(row=sheet.max_row, column=7, value=Re1)
          sheet.cell(row=sheet.max_row, column=8, value=S1)
          sheet.cell(row=sheet.max_row, column=9, value=fathername)
          sheet.cell(row=sheet.max_row, column=10, value=mothername)
          sheet.cell(row=sheet.max_row, column=11, value=F1)
          sheet.cell(row=sheet.max_row, column=12, value=M1)

      
          file.save('Students_data.xlsx')
       
          try:
            img.save("Student Images/"+str(R1)+".jpg")
          except:
               messagebox.showinfo("info","Profile picture is not available!!!!!!!!!")
          messagebox.showinfo("Success", "Data saved successfully.")
          clear()
# Function to search for data by registration number
def search_data(*args):
    text=Search.get()
    clear()
    saveButton.configure(state="disable")
    file=openpyxl.load_workbook('Student_data.xlsx')
    sheet=file.active
    
    for row in sheet.rows:
        if row[0].value ==int(text):
            name=row[0]
            reg_no_position=str(name)[14:-1]
            reg_number=str(name)[15:-1]
    try:
        print(str(name))
    except:
        messagebox.showerror("Invalid", "Invalid registration number!!!")
    x1=sheet.cell(row=int(reg_number),column=1).value
    x2=sheet.cell(row=int(reg_number),column=2).value
    x3=sheet.cell(row=int(reg_number),column=3).value
    x4=sheet.cell(row=int(reg_number),column=4).value
    x5=sheet.cell(row=int(reg_number),column=5).value
    x6=sheet.cell(row=int(reg_number),column=6).value
    x7=sheet.cell(row=int(reg_number),column=7).value
    x8=sheet.cell(row=int(reg_number),column=8).value
    x9=sheet.cell(row=int(reg_number),column=9).value
    x10=sheet.cell(row=int(reg_number),column=10).value
    x11=sheet.cell(row=int(reg_number),column=11).value
    x12=sheet.cell(row=int(reg_number),column=12).value
    
    
    Registration.set(x1)
    Name.set(x2)
    Course.set(x3)
    
    if x4=='Female':
        R2.select()
    else:
        R1.select()
    DOB.set(x5)    
    Date.set(x6)    
    Religion.set(x7)    
    CNIC.set(x8)    
    F_Name.set(x9)    
    M_Name.set(x10)    
    Father_Occupation.set(x11)    
    Mother_Occupation.set(x12) 
    
    img=(Image.open("Student Images/str(x1)+".jpg"))
    resized_image=img.resize((190,190))
    photo2 = ImageTk.PhotoImage(resized_image)
    lbl.configure(image=photo2)
    lbl.image=photo2          
                  
################################################
def update():
    R1=Registration.get()
    N1=Name.get()
    C1=Course.get()
    selection()
    G1=gender
    D2=DOB.get()
    D1=Date.get()
    Re1=Religion.get()
    S1=CNIC.get()
    fathername=F_Name.get()
    mothername=M_Name.get()
    F1=Father_Occupation.get()
    M1=Mother_Occupation.get()
    
    file=openpyxl.load_workbook('Student_data.xlsx')
    sheet=file.active
    for row in sheet.rows:
        if row[0].value ==R1
            name=row[0]
            reg_no_position=str(name)[14:-1]
            reg_number=str(name)[15:-1]
    
    sheet.cell(collumn=1,row=int(reg_number),value=R1)
    sheet.cell(collumn=2,row=int(reg_number),value=N1)
    sheet.cell(collumn=3,row=int(reg_number),value=C1)
    sheet.cell(collumn=4,row=int(reg_number),value=G1)
    sheet.cell(collumn=5,row=int(reg_number),value=D2)
    sheet.cell(collumn=6,row=int(reg_number),value=D1)
    sheet.cell(collumn=7,row=int(reg_number),value=Re1)
    sheet.cell(collumn=8,row=int(reg_number),value=S1)
    sheet.cell(collumn=9,row=int(reg_number),value=fathername)
    sheet.cell(collumn=10,row=int(reg_number),value=mothername)
    sheet.cell(collumn=11,row=int(reg_number),value=F1
    sheet.cell(collumn=12,row=int(reg_number),value=M1)
    
    file.save(r'Student_data.xlsx')
    
    try:
        img.save("Student Images/"+str(x1)+".jpg")
    except:
        pass
    messagebox.showinfo("Show info,"Update successfull")
    
    clear()
    

# Function to clear all entry fields
def clear_data(*args):
    global undo_history
    global img
    # Save the current state to undo history
    undo_history.append(get_all_entry_values())
    
    # Clear all entry fields
    #Registration.set("")
    Date.set(d1)
    Name.set("")
    DOB.set("")
    Gender.set(0)
    Religion.set("")
    CNIC.set("")
    Course.set("Select Course")
    F_Name.set("")
    Father_Occupation.set("")
    M_Name.set("")
    Mother_Occupation.set("")
   
    Registration_no()
    saveButton.configure(state="normal")


    img1=PhotoImage(file='Images/upload photo.png')
    lbl.configure(image=img1)
    lbl.image=img1
   
    img=""
    # Clear the image label
    img = Image.open("examples/Images/upload photo.png")
    resized_image = img.resize((190, 190))
    photo2 = ImageTk.PhotoImage(resized_image)
    lbl.config(image=photo2)
    lbl.image = photo2

    # Show success message
    messagebox.showinfo("Success", "Data cleared successfully.")
#################################################################################
# Function to get all entry values as a tuple
def get_all_entry_values():
    return (
        Registration.get(), Name.get(), DOB.get(), Gender.get(), Course.get(),
        Religion.get(), CNIC.get(), F_Name.get(), Father_Occupation.get(),
        M_Name.get(), Mother_Occupation.get()
    )

# Function to undo the last change
def undo(*args):
    global undo_history, redo_history
    if undo_history:
        # Save the current state to redo history
        redo_history.append(get_all_entry_values())
        
        # Restore the previous state
        previous_state = undo_history.pop()
        set_all_entry_values(previous_state)
        messagebox.showinfo("Undo", "Undo successful.")
    else:
        messagebox.showinfo("Undo", "No more undo steps.")

# Function to redo the last undone change
def redo(*args):
    global undo_history, redo_history
    if redo_history:
        # Save the current state to undo history
        undo_history.append(get_all_entry_values())
        
        # Restore the next state
        next_state = redo_history.pop()
        set_all_entry_values(next_state)
        messagebox.showinfo("Redo", "Redo successful.")
    else:
        messagebox.showinfo("Redo", "No more redo steps.")

# Function to set all entry values based on a tuple
def set_all_entry_values(entry_values):
    Registration.set(entry_values[0])
    Name.set(entry_values[1])
    DOB.set(entry_values[2])
    Gender.set(entry_values[3])
    Course.set(entry_values[4])
    Religion.set(entry_values[5])
    CNIC.set(entry_values[6])
    F_Name.set(entry_values[7])
    Father_Occupation.set(entry_values[8])
    M_Name.set(entry_values[9])
    Mother_Occupation.set(entry_values[10])

# Validation for Date of Birth
def validate_dob():
    dob = DOB.get()
    try:
        # Attempt to parse the date
        datetime.datetime.strptime(dob, '%d/%m/%Y')
        return True
    except ValueError:
        messagebox.showerror("Invalid Date", "Please enter a valid date in the format DD/MM/YYYY.")
        return False

# Validation for Email
def validate_email():
    email = "FA23-BCS-116@cuilahore.edu.pk"  # You can use the actual email variable here
    if not re.match(r"[^@]+@[^@]+\.[^@]+", email):
        messagebox.showerror("Invalid Email", "Please enter a valid email address.")
        return False
    return True

#top frames
CTkLabel(root, text="Email: FA23-BCS-116@cuilahore.edu.pk", width=100, height=30, bg_color="#f0687c", anchor='e').pack(side=TOP, fill=X,padx=5)
CTkLabel(root, text="STUDENT REGISTRATION", width=100, height=20, bg_color="#c36464", fg_color="#fff", font=("Arial" ,20 ,"bold")).pack(side="top", fill=X)

#search box to update
Search = StringVar()
CTkEntry(root, textvariable=Search, width=150, font=("Arial", 20)).place(x=830, y=85)
Srch = CTkButton(root, text="Search \U0001F50E", compound=LEFT, width=120,height=50, bg_color="#68ddfa", font=("Arial" ,13, "bold"),command=search_data)
Srch.place(x=1030, y=70.5)
root.bind("<Control-Key-s>",search_data)
imageicon4 = PhotoImage(file="Images/Layer 4.png")
Update_button = Button(root, image=imageicon4, bg="#c36464",command=Update).place(x=110, y=64)

#Registration and Date
Label(root, text="Registration No.", font="Arial 13", fg=framebg, bg=background).place(x=30, y=150)
CTkLabel(root, text="Date: ", font=("Arial", 13), fg_color=framebg, bg_color=background).place(x=500, y=150)

Registration=StringVar()
Date = StringVar()

reg_entry = CTkEntry(root, textvariable=Registration, width=150, font=("Arial", 22))
reg_entry.place(x=160, y=150)

registration_no()
today = date.today()
d1 = today.strftime("%d/%m/%Y")
date_entry = CTkEntry(root, textvariable=Date, width=130, font=("Arial", 18)).place(x=540, y=150)

Date.set(d1)

#Student details
obj = LabelFrame(root, text="Student's Details", font=20, bd=2, width=900, bg=framebg, fg=framefg, height=250, relief=GROOVE)
obj.place(x=30, y=200)

Label(obj, text="Full Name:", font=("Arial", 13), bg=framebg, fg=framefg).place(x=30, y=50)
Label(obj, text="Date of Birth:", font="Arial 13", bg=framebg, fg=framefg).place(x=30, y=100)
Label(obj, text="Gender:", font="Arial 13", bg=framebg, fg=framefg).place(x=30, y=150)

Label(obj, text="Course:", font="Arial 13", bg=framebg, fg=framefg).place(x=500, y=50)
Label(obj, text="Religion:", font="Arial 13", bg=framebg, fg=framefg).place(x=500, y=100)
Label(obj, text="CNIC:", font="Arial 13", bg=framebg, fg=framefg).place(x=500, y=150)

Name=StringVar()
name_entry = CTkEntry(obj, textvariable=Name, width=190, font=("Arial", 16)).place(x=160, y=50)

DOB=StringVar()
dob_entry = CTkEntry(obj, textvariable=DOB, width=190, font=("Arial", 16)).place(x=160, y=100)

Gender=StringVar()
R1=CTkRadioButton(obj, text="Male", variable=Gender, value=1, bg_color=framebg, fg_color=framefg, command=selection).place(x=150, y=150)
R2=CTkRadioButton(obj, text="Female", variable=Gender, value=2, bg_color=framebg, fg_color=framefg, command=selection).place(x=220, y=150)
root.bind("<Control-Key-g>",selection)
Religion=StringVar()
religion_entry = CTkEntry(obj, textvariable=Religion, width=180, font=("Arial", 16)).place(x=630, y=100)

CNIC=StringVar()
CNIC_entry = CTkEntry(obj, textvariable=CNIC, width=190, font=("Arial", 16)).place(x=630, y=150)
c1=StringVar()

Course=CTkComboBox(obj,values=['CS','SE','CE','BF','AI','PH','EE','ME','CHE','MATHS','Stats','Ds','ENG','PHY'], font=("Arial", 18), width=150,text_color="blue",corner_radius=2,dropdown_fg_color="white",dropdown_text_color="red",dropdown_font=("helvetica",15),dropdown_hover_color="blue",variable=c1)
Course.place(x=630, y=50)
Course.set("Select Course")

#Parents details
obj2 = LabelFrame(root, text="Parent's Details", font=20, bd=2, width=900, bg=framebg, fg=framefg, height=220, relief=GROOVE)
obj2.place(x=30, y=470)

Label(obj2, text="Father's Name:", font="Arial 14", bg=framebg, fg=framefg).place(x=30, y=50)
Label(obj2, text="Occupation:", font="Arial 14", bg=framebg, fg=framefg).place(x=30, y=100)

F_Name=StringVar()
f_entry = CTkEntry(obj2,textvariable=F_Name, width=190, font=("Arial", 16))
f_entry.place(x=160, y=50)

Father_Occupation=StringVar()
F0_entry = CTkEntry(obj2, textvariable=Father_Occupation, width=190, font=("Arial", 16))
F0_entry.place(x=160, y=100)

Label(obj2, text="Mother's Name:", font="Arial 14", bg=framebg, fg=framefg).place(x=500, y=50)
Label(obj2, text="Occupation:", font="Arial 14", bg=framebg, fg=framefg).place(x=500, y=100)

M_Name=StringVar()
m_entry = CTkEntry(obj2,textvariable=M_Name, width=190, font=("Arial", 16))
m_entry.place(x=630, y=50)

Mother_Occupation=StringVar()
M0_entry = CTkEntry(obj2, textvariable=Mother_Occupation, width=190, font=("Arial", 16))
M0_entry.place(x=630, y=100)

#image
f=CTkFrame(root, border_width=2, bg_color="Black", width=195, height=195)
f.place(x=1000, y=140)
img=CTkImage(Image.open("examples/Images/upload photo.png"),size=(216,222))
lbl=CTkLabel(f, bg_color="Black", image=img,text="")
lbl.place(x=0, y=0)

#button
CTkButton(root, text="Upload", width=190, height=50, font=("Arial", 14,"bold"), bg_color="lightblue", command=showimage,hover_color="grey").place(x=1000, y=360)
root.bind("<Control-Key-u>",showimage)
saveButton = CTkButton(root, text="Save", width=190, height=50, font=("Arial" ,14, "bold"), bg_color="lightgreen",command=save_data,hover_color="green")
saveButton.place(x=1000, y=495)
root.bind("<Button-3>",save_data)

# CTkButton(root, text="Reset", width=190, height=50, font=("Arial", 12, "bold"), bg_color="lightpink",command=reset_form).place(x=1000, y=580)
# root.bind("<Control-Key-r>",reset_form)
CTkButton(root, text="Exit",   width=190, height=50, font=("Arial", 14, "bold"), bg_color="grey", command=Exit,hover_color="red").place(x=1000, y=620)
root.bind("<Control-Key-Escape>",Exit)
clearButton = CTkButton(root, text="Clear Data",  width=190, height=50, font=("Arial", 14, "bold"), bg_color="lightblue", command=clear_data,cursor="hand2",hover_color="white")
clearButton.place(x=680, y=150)
root.bind("<Control-Key-c>",clear_data)

# Undo and Redo buttons
undoButton = CTkButton(root, text="Undo", width=190, height=50, font=("Arial", 14, "bold"), bg_color="lightgray", command=undo,hover_color="grey")
undoButton.place(x=1000, y=425)
root.bind("<Control-Key-z>",undo)
redoButton = CTkButton(root, text="Redo", width=190, height=50, font=("Arial", 14, "bold"), bg_color="lightgray", command=redo,hover_color="grey")
redoButton.place(x=1000, y=555)
root.bind("<Control-Key-r>",redo)
root.mainloop()
